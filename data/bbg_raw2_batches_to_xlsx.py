#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sat Apr 10 05:53:03 2021

@author: ms
"""
import os
import tqdm
import platform
import numpy as np
import pandas as pd
from openpyxl import load_workbook
pd.set_option('display.max_columns', None)

if platform.node() in ['mstp3', 'msbq']:
    os.chdir('/home/ms/github/fxpred/data')

fp_raw = 'bbg_raw'
fp_clean = 'bbg'

if not os.path.isdir(fp_clean):
    os.makedirs(fp_clean)

for fp in sorted([i for i in os.listdir(fp_raw + '2') if 'batch' in i]):
    print(fp)
    workbook = load_workbook(os.path.join(fp_raw + '2', fp))
    for i in tqdm.tqdm(workbook.sheetnames):
        x = pd.read_excel(os.path.join(fp_raw + '2', fp), sheet_name=i, header=None)
        x = x.loc[:, ~x.isna().all(axis=0)]
        x.to_excel(
            os.path.join(
                fp_raw + '2',
                i.lower() + '.xlsx'
            ),
            header=False,
            index=False
        )
